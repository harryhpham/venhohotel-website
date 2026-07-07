"""Phase 3 Discovery v3 - tim nut Xuat file sau khi dong dialog"""
import sys, io, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

from playwright.sync_api import sync_playwright

LOGIN_URL = 'https://admin1.skyhotel.vn/login.aspx'

USER = os.environ.get('SKYHOTEL_USER', '')
PASS = os.environ.get('SKYHOTEL_PASS', '')

with sync_playwright() as p:
    browser = p.chromium.launch(headless=True, args=['--no-sandbox', '--disable-dev-shm-usage'])
    page = browser.new_context().new_page()

    # Login
    print('[1] Dang nhap...')
    page.goto(LOGIN_URL, wait_until='networkidle')
    page.fill('#txt_username', USER)
    page.fill('#txt_password', PASS)
    page.click('#cmd_login')
    page.wait_for_load_state('networkidle')
    print(f'    OK - URL: {page.url}')

    # Click menu
    print('[2] Navigate to Doanh thu hoa don...')
    for header in page.query_selector_all('li.menu_sub_1 > a'):
        text = header.inner_text().strip()
        if 'doanh thu' in text.lower():
            header.click(force=True)
            page.wait_for_timeout(1000)
            break
    page.evaluate("document.querySelector('a[href=\"#revenue_invoices\"]').click()")
    page.wait_for_timeout(5000)  # doi AJAX content load
    print(f'    URL: {page.url}')

    # Fill ngay vao date picker va click OK de load du lieu
    print('[3] Fill ngay va submit date picker...')
    import datetime
    yesterday = (datetime.date.today() - datetime.timedelta(days=1)).strftime('%d/%m/%Y')
    print(f'    Ngay: {yesterday}')

    # Fill bang JavaScript (co the readonly qua UI)
    page.evaluate(f"""
        var d = '{yesterday}';
        var dbegin = document.getElementById('date_begin');
        var dend = document.getElementById('date_end');
        if (dbegin) dbegin.value = d;
        if (dend) dend.value = d;
    """)
    page.wait_for_timeout(500)

    # Click OK button cua fancy date picker
    ok_btn = page.query_selector('#fancyConfirmdate_edit')
    if ok_btn:
        ok_btn.click(force=True)
        print('    Da click OK - doi AJAX load...')
        page.wait_for_timeout(5000)
    else:
        print('    Khong tim thay OK button!')

    page.screenshot(path='AI Agent/discovery-revenue.png', full_page=True)
    print('    Screenshot: AI Agent/discovery-revenue.png')

    # Dump tat ca links va buttons sau khi data load
    print('[4] Tat ca buttons/links sau khi data load:')
    for el in page.query_selector_all('button, input[type=submit], input[type=button], input[type=image], a[onclick], a[class*=btn]'):
        id_ = el.get_attribute('id') or ''
        cls = (el.get_attribute('class') or '')[:60]
        onclick = (el.get_attribute('onclick') or '')[:100]
        href = el.get_attribute('href') or ''
        try: text = el.inner_text().strip()[:60]
        except: text = ''
        if id_ or text or onclick:
            print(f'    id="{id_}"  text="{text}"  onclick="{onclick}"  href="{href}"  class="{cls}"')

    # Tim element co text lien quan export
    print('[5] Elements co text Xuat/Export/Excel/Tien/File:')
    for kw in ['Xuất', 'Export', 'Excel', 'Tải', 'File', 'xuat', 'In ', 'Print']:
        for el in page.get_by_text(kw, exact=False).all()[:5]:
            try:
                tag = el.evaluate("e=>e.tagName")
                el_id = el.get_attribute('id') or ''
                el_cls = (el.get_attribute('class') or '')[:60]
                el_onclick = (el.get_attribute('onclick') or '')[:100]
                el_text = el.inner_text().strip()[:60]
                if tag in ['A', 'BUTTON', 'INPUT', 'SPAN', 'DIV']:
                    print(f'    "{el_text}"  <{tag}>  id="{el_id}"  onclick="{el_onclick}"  class="{el_cls}"')
            except:
                pass

    # Download file Excel va xem columns
    print('[6] Download file Excel...')
    import os
    save_path = os.path.join(os.path.dirname(__file__), 'test-export.xlsx')
    try:
        with page.expect_download(timeout=30000) as dl_info:
            page.click('#export_revenue_v1')
        download = dl_info.value
        download.save_as(save_path)
        print(f'    Saved: {save_path}')
        print(f'    Filename: {download.suggested_filename}')

        # Doc file Excel
        import openpyxl
        wb = openpyxl.load_workbook(save_path)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        print(f'\n    === Excel: {len(rows)} rows, {ws.max_column} cols ===')
        print(f'    Headers: {rows[0]}')
        print(f'\n    Row 1: {rows[1] if len(rows) > 1 else "(empty)"}')
        print(f'    Row 2: {rows[2] if len(rows) > 2 else "(empty)"}')
        print(f'    Row 3: {rows[3] if len(rows) > 3 else "(empty)"}')
    except Exception as e:
        print(f'    ERROR: {e}')
        import traceback
        traceback.print_exc()

    browser.close()

print('\nDONE')

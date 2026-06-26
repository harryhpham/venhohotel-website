"""
SkyHotel Scraper -- Ven Ho Hotel
Chạy hàng ngày (doanh thu + phiếu chi) và tổng kết tháng.
"""

import sys
import os
import datetime
import calendar
import smtplib
import openpyxl
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from playwright.sync_api import sync_playwright

# --- CONFIG -------------------------------------------------------------------
LOGIN_URL      = "https://admin1.skyhotel.vn/login.aspx"

SKYHOTEL_USER  = os.environ.get("SKYHOTEL_USER", "")
SKYHOTEL_PASS  = os.environ.get("SKYHOTEL_PASS", "")
GMAIL_USER     = os.environ.get("GMAIL_USER", "")
GMAIL_APP_PASS = os.environ.get("GMAIL_APP_PASS", "")

# Selectors — Revenue
SEL_USERNAME       = "#txt_username"
SEL_PASSWORD       = "#txt_password"
SEL_LOGIN_BTN      = "#cmd_login"
SEL_DATE_OK        = "#fancyConfirmdate_edit"
SEL_EXPORT_REVENUE = "#export_revenue_v1"

# Selectors — Phiếu Chi
SEL_PC_MIN    = "#min_time"
SEL_PC_MAX    = "#max_time"
SEL_PC_RELOAD = "#cost_reload"
SEL_PC_EXPORT = "#cost_list_export"

# Revenue Excel columns (data starts row index 8, 0-indexed)
COL_STT        = 0
COL_PHONG      = 2
COL_KHACH      = 3
COL_TIEN_PHONG = 7
COL_TIEN_DV    = 8
COL_TONG_CONG  = 11
COL_HINH_THUC  = 16

# Phiếu Chi Excel columns (data starts row index 8, 0-indexed)
PC_STT      = 0
PC_SO_CT    = 1
PC_NGAY_CHI = 2
PC_DIEN_GIAI = 3
PC_SO_TIEN  = 4
PC_HTTT     = 5
PC_LOAI     = 6
PC_NGUOI_NHAN = 7


# --- LOGIN --------------------------------------------------------------------

def do_login(page):
    page.goto(LOGIN_URL, wait_until="networkidle", timeout=30000)
    page.fill(SEL_USERNAME, SKYHOTEL_USER)
    page.fill(SEL_PASSWORD, SKYHOTEL_PASS)
    page.click(SEL_LOGIN_BTN)
    page.wait_for_load_state("networkidle")
    return "login" not in page.url.lower()


# --- NAVIGATE -----------------------------------------------------------------

def navigate_to_revenue(page, date_from: str, date_to: str = None):
    """date_from / date_to format: 'dd/mm/yyyy'"""
    if date_to is None:
        date_to = date_from
    for header in page.query_selector_all("li.menu_sub_1 > a"):
        if "doanh thu" in header.inner_text().lower():
            header.click(force=True)
            page.wait_for_timeout(800)
            break
    page.evaluate("document.querySelector('a[href=\"#revenue_invoices\"]').click()")
    page.wait_for_timeout(4000)
    page.evaluate(f"""
        var b = document.getElementById('date_begin');
        var e = document.getElementById('date_end');
        if (b) b.value = '{date_from}';
        if (e) e.value = '{date_to}';
    """)
    page.wait_for_timeout(300)
    ok = page.query_selector(SEL_DATE_OK)
    if ok:
        ok.click(force=True)
    page.wait_for_timeout(5000)


def navigate_to_phieu_chi(page, date_from: str = None, date_to: str = None):
    for header in page.query_selector_all("li.menu_sub_1 > a"):
        if "tiền chi" in header.inner_text().lower():
            header.click(force=True)
            page.wait_for_timeout(800)
            break
    page.evaluate("document.querySelector('a[href=\"#cost_list\"]').click()")
    page.wait_for_timeout(4000)

    if date_from and date_to:
        try:
            page.fill(SEL_PC_MIN, date_from)
            page.fill(SEL_PC_MAX, date_to)
            page.wait_for_timeout(300)
            page.click(SEL_PC_RELOAD)
            page.wait_for_timeout(3000)
        except Exception:
            pass


# --- DOWNLOAD -----------------------------------------------------------------

def download_file(page, export_selector: str, save_path: str) -> str:
    with page.expect_download(timeout=30000) as dl_info:
        page.click(export_selector)
    dl_info.value.save_as(save_path)
    return save_path


# --- PARSE --------------------------------------------------------------------

def parse_revenue(filepath: str, label: str) -> dict:
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    def is_data_row(r):
        stt = r[COL_STT]
        if stt is None:
            return False
        if isinstance(stt, (int, float)):
            return True
        return isinstance(stt, str) and stt.strip().isdigit()

    data_rows = [r for r in rows[8:] if is_data_row(r)]

    total_revenue = 0
    total_room    = 0
    total_svc     = 0
    rooms_set     = set()
    payment_methods = {}
    room_revenue  = {}

    for r in data_rows:
        tong      = r[COL_TONG_CONG] or 0
        tien_phong = r[COL_TIEN_PHONG] or 0
        tien_dv   = r[COL_TIEN_DV] or 0

        if isinstance(tong, (int, float)):      total_revenue += tong
        if isinstance(tien_phong, (int, float)): total_room   += tien_phong
        if isinstance(tien_dv, (int, float)):    total_svc    += tien_dv

        phong = str(r[COL_PHONG] or "").strip()
        if phong:
            rooms_set.add(phong)
            room_revenue[phong] = room_revenue.get(phong, 0) + (tong if isinstance(tong, (int, float)) else 0)

        httt = str(r[COL_HINH_THUC] or "").strip()
        if httt:
            payment_methods[httt] = payment_methods.get(httt, 0) + 1

    top_rooms = sorted(room_revenue.items(), key=lambda x: x[1], reverse=True)[:5]
    guests = [str(r[COL_KHACH] or "").strip() for r in data_rows if r[COL_KHACH]]

    return {
        "label": label,
        "total_invoices": len(data_rows),
        "total_revenue":  int(total_revenue),
        "room_revenue":   int(total_room),
        "service_revenue": int(total_svc),
        "rooms_checked_out": sorted(rooms_set),
        "rooms_count":    len(rooms_set),
        "payment_methods": payment_methods,
        "top_rooms":      top_rooms,
        "guests":         guests,
    }


def parse_phieu_chi(filepath: str, filter_date: str = None, filter_month: str = None) -> dict:
    """
    filter_date:  'dd/mm/yyyy' — lọc theo ngày cụ thể (daily report)
    filter_month: 'mm/yyyy'   — lọc theo tháng (monthly report)
    Không truyền → lấy tất cả rows trong file.
    """
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    def is_data_row(r):
        stt = r[PC_STT]
        if stt is None:
            return False
        if isinstance(stt, (int, float)):
            return True
        return isinstance(stt, str) and stt.strip().isdigit()

    data_rows = [r for r in rows[8:] if is_data_row(r)]

    if filter_date:
        data_rows = [r for r in data_rows if str(r[PC_NGAY_CHI] or "").startswith(filter_date)]
    elif filter_month:
        # filter_month = 'mm/yyyy', ngay chi = 'dd/mm/yyyy HH:MM'
        data_rows = [r for r in data_rows if f"/{filter_month}" in str(r[PC_NGAY_CHI] or "")]

    items = []
    total_expense = 0
    by_category   = {}

    for r in data_rows:
        so_tien = r[PC_SO_TIEN] or 0
        if isinstance(so_tien, (int, float)):
            total_expense += so_tien
        loai = str(r[PC_LOAI] or "").strip()
        if loai and isinstance(so_tien, (int, float)):
            by_category[loai] = by_category.get(loai, 0) + so_tien

        items.append({
            "so_ct":     str(r[PC_SO_CT] or "").strip(),
            "ngay_chi":  str(r[PC_NGAY_CHI] or "").strip(),
            "dien_giai": str(r[PC_DIEN_GIAI] or "").strip(),
            "so_tien":   int(so_tien) if isinstance(so_tien, (int, float)) else 0,
            "httt":      str(r[PC_HTTT] or "").strip(),
            "loai":      loai,
        })

    return {
        "items":         items,
        "total_expense": int(total_expense),
        "total_count":   len(items),
        "by_category":   by_category,
    }


# --- FORMAT -------------------------------------------------------------------

def vnd(amount: int) -> str:
    return f"{amount:,}đ".replace(",", ".")


def build_daily_email(revenue: dict, phieu_chi: dict) -> str:
    profit = revenue["total_revenue"] - phieu_chi["total_expense"]
    rooms_str = ", ".join(revenue["rooms_checked_out"]) or "(không có)"

    httt_lines = "\n".join(
        f"  • {k}: {v} hóa đơn" for k, v in revenue["payment_methods"].items()
    ) or "  (không có)"

    top_str = "\n".join(
        f"  • Phòng {r}: {vnd(int(v))}" for r, v in revenue["top_rooms"]
    ) or "  (không có)"

    if phieu_chi["items"]:
        pc_lines = "\n".join(
            f"  • {it['so_ct']} | {it['loai'] or it['dien_giai']} | {vnd(it['so_tien'])} ({it['httt']})"
            for it in phieu_chi["items"]
        )
        pc_cat = "\n".join(
            f"  • {k}: {vnd(int(v))}" for k, v in phieu_chi["by_category"].items()
        )
    else:
        pc_lines = "  (không có phiếu chi)"
        pc_cat   = "  (không có)"

    return f"""BÁO CÁO DOANH THU — VEN HỒ HOTEL
Ngày: {revenue["label"]}
Nguồn: SkyHotel PMS (admin1.skyhotel.vn)

=== TỔNG QUAN ===
• Tổng doanh thu:      {vnd(revenue["total_revenue"])}
• Tiền phòng:          {vnd(revenue["room_revenue"])}
• Dịch vụ:             {vnd(revenue["service_revenue"])}
• Tổng chi phí:        {vnd(phieu_chi["total_expense"])}  ({phieu_chi["total_count"]} phiếu)
• Lợi nhuận ước tính:  {vnd(profit)}
• Hóa đơn checkout:    {revenue["total_invoices"]}
• Phòng checkout:      {revenue["rooms_count"]}/12
• Phòng: {rooms_str}

=== HÌNH THỨC THANH TOÁN (Thu) ===
{httt_lines}

=== TOP PHÒNG DOANH THU CAO ===
{top_str}

=== PHIẾU CHI ({phieu_chi["total_count"]} phiếu — {vnd(phieu_chi["total_expense"])}) ===
{pc_lines}

  Chi theo loại:
{pc_cat}

---
Email tự động từ AI Agent lúc {datetime.datetime.now().strftime("%H:%M %d/%m/%Y")}
"""


def build_monthly_email(revenue: dict, phieu_chi: dict, month_label: str) -> str:
    profit = revenue["total_revenue"] - phieu_chi["total_expense"]

    httt_lines = "\n".join(
        f"  • {k}: {v} hóa đơn" for k, v in revenue["payment_methods"].items()
    ) or "  (không có)"

    top_str = "\n".join(
        f"  • Phòng {r}: {vnd(int(v))}" for r, v in revenue["top_rooms"]
    ) or "  (không có)"

    pc_cat = "\n".join(
        f"  • {k}: {vnd(int(v))}" for k, v in phieu_chi["by_category"].items()
    ) or "  (không có)"

    return f"""BÁO CÁO THÁNG {month_label} — VEN HỒ HOTEL
Nguồn: SkyHotel PMS (admin1.skyhotel.vn)

=== TỔNG KẾT THÁNG ===
• Tổng doanh thu:      {vnd(revenue["total_revenue"])}
• Tiền phòng:          {vnd(revenue["room_revenue"])}
• Dịch vụ:             {vnd(revenue["service_revenue"])}
• Tổng chi phí:        {vnd(phieu_chi["total_expense"])}  ({phieu_chi["total_count"]} phiếu)
• Lợi nhuận ước tính:  {vnd(profit)}
• Tổng hóa đơn:        {revenue["total_invoices"]}
• Tổng phòng checkout: {revenue["rooms_count"]}

=== HÌNH THỨC THANH TOÁN ===
{httt_lines}

=== TOP PHÒNG DOANH THU CAO (cả tháng) ===
{top_str}

=== CHI PHÍ THEO LOẠI ===
{pc_cat}

---
Email tự động từ AI Agent lúc {datetime.datetime.now().strftime("%H:%M %d/%m/%Y")}
"""


# --- EMAIL --------------------------------------------------------------------

def send_email(subject: str, body: str) -> bool:
    if not GMAIL_USER or not GMAIL_APP_PASS:
        print("[Email] GMAIL_USER / GMAIL_APP_PASS chưa set — bỏ qua")
        return False
    msg = MIMEMultipart()
    msg["From"]    = GMAIL_USER
    msg["To"]      = GMAIL_USER
    msg["Subject"] = subject
    msg.attach(MIMEText(body, "plain", "utf-8"))
    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as srv:
            srv.login(GMAIL_USER, GMAIL_APP_PASS)
            srv.send_message(msg)
        print(f"[Email] Đã gửi: {subject}")
        return True
    except Exception as e:
        print(f"[Email] Lỗi: {e}")
        return False


# --- MAIN FLOWS ---------------------------------------------------------------

def run_daily(report_date: str = None, download_dir: str = None) -> dict:
    if not report_date:
        report_date = (datetime.date.today() - datetime.timedelta(days=1)).strftime("%d/%m/%Y")
    if not download_dir:
        download_dir = os.path.dirname(__file__) or "."
    if not SKYHOTEL_USER or not SKYHOTEL_PASS:
        raise ValueError("Thiếu SKYHOTEL_USER / SKYHOTEL_PASS")

    rev_path = os.path.join(download_dir, "skyhotel_report.xlsx")
    pc_path  = os.path.join(download_dir, "phieu_chi_report.xlsx")

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True, args=["--no-sandbox", "--disable-dev-shm-usage"])
        page    = browser.new_context().new_page()

        if not do_login(page):
            raise RuntimeError("Đăng nhập SkyHotel thất bại")

        navigate_to_revenue(page, report_date)
        download_file(page, SEL_EXPORT_REVENUE, rev_path)

        navigate_to_phieu_chi(page)
        download_file(page, SEL_PC_EXPORT, pc_path)

        browser.close()

    revenue   = parse_revenue(rev_path, report_date)
    phieu_chi = parse_phieu_chi(pc_path, filter_date=report_date)

    body    = build_daily_email(revenue, phieu_chi)
    subject = f"[Ven Hồ] Báo cáo doanh thu {report_date}"
    send_email(subject, body)
    return {"revenue": revenue, "phieu_chi": phieu_chi, "email_body": body}


def run_monthly(year: int = None, month: int = None, download_dir: str = None) -> dict:
    today = datetime.date.today()
    if year is None or month is None:
        first_of_this_month = today.replace(day=1)
        prev   = first_of_this_month - datetime.timedelta(days=1)
        year, month = prev.year, prev.month

    if not download_dir:
        download_dir = os.path.dirname(__file__) or "."
    if not SKYHOTEL_USER or not SKYHOTEL_PASS:
        raise ValueError("Thiếu SKYHOTEL_USER / SKYHOTEL_PASS")

    last_day    = calendar.monthrange(year, month)[1]
    date_from   = f"01/{month:02d}/{year}"
    date_to     = f"{last_day:02d}/{month:02d}/{year}"
    month_label = f"{month:02d}/{year}"
    filter_month = f"{month:02d}/{year}"

    rev_path = os.path.join(download_dir, "skyhotel_monthly.xlsx")
    pc_path  = os.path.join(download_dir, "phieu_chi_monthly.xlsx")

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True, args=["--no-sandbox", "--disable-dev-shm-usage"])
        page    = browser.new_context().new_page()

        if not do_login(page):
            raise RuntimeError("Đăng nhập SkyHotel thất bại")

        navigate_to_revenue(page, date_from, date_to)
        download_file(page, SEL_EXPORT_REVENUE, rev_path)

        navigate_to_phieu_chi(page, date_from, date_to)
        download_file(page, SEL_PC_EXPORT, pc_path)

        browser.close()

    revenue   = parse_revenue(rev_path, f"Tháng {month_label} ({date_from} → {date_to})")
    phieu_chi = parse_phieu_chi(pc_path, filter_month=filter_month)

    body    = build_monthly_email(revenue, phieu_chi, month_label)
    subject = f"[Ven Hồ] Báo cáo tháng {month_label}"
    send_email(subject, body)
    return {"revenue": revenue, "phieu_chi": phieu_chi, "email_body": body}


# --- ENTRY POINT --------------------------------------------------------------

if __name__ == "__main__":
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

    if "--monthly" in sys.argv:
        result = run_monthly()
    else:
        report_date = next((a for a in sys.argv[1:] if not a.startswith("--")), None)
        result = run_daily(report_date)

    print(result["email_body"])
    if "--monthly" not in sys.argv:
        print("\nDữ liệu raw:")
        for k, v in result["revenue"].items():
            if k != "guests":
                print(f"  {k}: {v}")
        print(f"  phieu_chi total: {result['phieu_chi']['total_expense']}")
